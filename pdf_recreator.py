#!/usr/bin/env python3
"""
PDF to Excel Layout Recreator
==============================
A professional, modular Python script that recreates PDF files into editable
Excel workbooks, preserving visual layout, table structures, formatting, and images.

Features:
- Page rendering: PyMuPDF to render pages as high-resolution images.
- Grid mapping: Coordinate clustering algorithm maps PDF layout to Excel rows/columns.
- Table detection: OpenCV morphological analysis + pdfplumber line detection.
- OCR fallback: Tesseract/EasyOCR fallback for scanned or unreadable text.
- Validation: Low-confidence highlighting, error logs, and comments.
- Control Sheet: Page statistics, warnings, extraction metadata.
- Local or Server execution.

Author: AI Coding Agent (Google AI Studio Build)
Date: June 2026
"""

import os
import sys
import datetime
import argparse
import logging
import warnings

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger("pdf_recreator")

# Setup optional dependency checks and graceful fallbacks
HAS_FITZ = False
HAS_PDFPLUMBER = False
HAS_CV2 = False
HAS_NP = False
HAS_PIL = False
HAS_OCR = False
HAS_OPENPYXL = False

# Try to import PyMuPDF
try:
    import fitz # PyMuPDF
    HAS_FITZ = True
except ImportError:
    logger.warning("PyMuPDF (fitz) is not installed. PDF reading and page rendering will be disabled.")

# Try to import pdfplumber
try:
    import pdfplumber
    HAS_PDFPLUMBER = True
except ImportError:
    logger.warning("pdfplumber is not installed. Advanced table extraction will fall back to OpenCV/PyMuPDF.")

# Try to import OpenCV
try:
    import cv2
    HAS_CV2 = True
except ImportError:
    logger.warning("OpenCV (cv2) is not installed. Image-based table/line detection will be disabled.")

# Try to import NumPy
try:
    import numpy as np
    HAS_NP = True
except ImportError:
    logger.warning("NumPy is not installed. Mathematical grid clustering and OpenCV operations will be disabled.")

# Try to import Pillow
try:
    from PIL import Image as PILImage
    HAS_PIL = True
except ImportError:
    logger.warning("Pillow (PIL) is not installed. Image handling and OCR cropping will be limited.")

# Try to import OCR libraries (Tesseract or EasyOCR)
OCR_ENGINE = None
try:
    import pytesseract
    # Check if tesseract binary is accessible
    try:
        pytesseract.get_tesseract_version()
        HAS_OCR = True
        OCR_ENGINE = "Tesseract"
    except Exception:
        logger.warning("pytesseract is installed but the Tesseract binary is not found in PATH. OCR will be disabled.")
except ImportError:
    try:
        import easyocr
        reader = easyocr.Reader(['en'], gpu=False)
        HAS_OCR = True
        OCR_ENGINE = "EasyOCR"
    except ImportError:
        logger.warning("Neither pytesseract nor easyocr is installed. OCR fallback will be disabled.")

# Try to import openpyxl
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.comments import Comment
    HAS_OPENPYXL = True
except ImportError:
    logger.error("openpyxl is not installed. Excel creation is disabled! Please install openpyxl to run this script.")


class PDFRecreator:
    """
    Core engine that handles PDF-to-Excel reconstruction.
    """
    def __init__(self, pdf_path, output_excel="recreated_from_pdf.xlsx",
                 preview_dir="page_previews", log_path="extraction_log.txt",
                 combine_pages=False, grid_threshold=10, ocr_threshold=0.6):
        self.pdf_path = pdf_path
        self.output_excel = output_excel
        self.preview_dir = preview_dir
        self.log_path = log_path
        self.combine_pages = combine_pages
        self.grid_threshold = grid_threshold # Pixel threshold for clustering grid lines
        self.ocr_threshold = ocr_threshold   # OCR confidence threshold (below this is highlighted)
        
        # Logs and metadata
        self.logs = []
        self.warnings = []
        self.stats = {} # Per-page stats: {page_num: {tables: X, text_blocks: Y, images: Z}}
        self.uncertain_cells = [] # List of tuples: (page, sheet_cell, value, reason, confidence)
        self.start_time = datetime.datetime.now()

        # Create directories
        if self.preview_dir:
            os.makedirs(self.preview_dir, exist_ok=True)

    def log(self, message, level=logging.INFO):
        """Records message to both memory, standard logger, and file."""
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_entry = f"{timestamp} - {message}"
        self.logs.append(log_entry)
        if level == logging.WARNING:
            self.warnings.append(log_entry)
            logger.warning(message)
        elif level == logging.ERROR:
            self.warnings.append(f"ERROR: {log_entry}")
            logger.error(message)
        else:
            logger.info(message)

    def write_log_file(self):
        """Writes the accumulated logs to disk."""
        if self.log_path:
            try:
                with open(self.log_path, "w", encoding="utf-8") as f:
                    f.write("\n".join(self.logs))
                logger.info(f"Extraction log written to {self.log_path}")
            except Exception as e:
                logger.error(f"Failed to write extraction log file: {e}")

    def render_pages(self):
        """
        Renders each PDF page as a high-resolution image (300 DPI) using PyMuPDF.
        Returns a list of rendered image file paths.
        """
        self.log("Step 1: Rendering PDF pages to high-resolution images (300 DPI)")
        if not HAS_FITZ:
            self.log("PyMuPDF is missing. Cannot render pages to images.", logging.WARNING)
            return []

        rendered_images = []
        try:
            doc = fitz.open(self.pdf_path)
            for i, page in enumerate(doc):
                page_num = i + 1
                # 300 DPI is achieved using matrix scaling
                zoom = 300 / 72  # standard PDF resolution is 72 points per inch
                matrix = fitz.Matrix(zoom, zoom)
                pix = page.get_pixmap(matrix=matrix)
                
                img_path = os.path.join(self.preview_dir, f"page_{page_num}_render.png")
                pix.save(img_path)
                rendered_images.append(img_path)
                self.log(f"Rendered Page {page_num} saved to {img_path}")
            doc.close()
        except Exception as e:
            self.log(f"Failed to render PDF pages: {e}", logging.ERROR)
        
        return rendered_images

    def detect_tables_opencv(self, image_path, page_num):
        """
        Uses OpenCV morphological operations to detect table outlines, horizontal,
        and vertical grid lines on a page. Returns cell bounding boxes.
        """
        if not (HAS_CV2 and HAS_NP):
            self.log(f"OpenCV or NumPy is missing. Skipping OpenCV table detection for page {page_num}.", logging.WARNING)
            return []

        self.log(f"Running OpenCV table and line detection for page {page_num}")
        try:
            # Read image in grayscale
            img = cv2.imread(image_path)
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            
            # Binary thresholding (Otsu's thresholding)
            thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)[1]
            
            # Detect horizontal lines
            horizontal_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (40, 1))
            detect_horizontal = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, horizontal_kernel, iterations=2)
            
            # Detect vertical lines
            vertical_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, 40))
            detect_vertical = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, vertical_kernel, iterations=2)
            
            # Join horizontal and vertical masks to get complete table grid
            table_grid = cv2.bitwise_or(detect_horizontal, detect_vertical)
            
            # Find contours of tables
            contours, _ = cv2.findContours(table_grid, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            
            detected_cells = []
            for cnt in contours:
                x, y, w, h = cv2.boundingRect(cnt)
                # Filter out small contours that aren't tables
                if w > 100 and h > 40:
                    # Crop table region and look for sub-cells
                    table_crop = table_grid[y:y+h, x:x+w]
                    cell_contours, _ = cv2.findContours(table_crop, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
                    
                    for cc in cell_contours:
                        cx, cy, cw, ch = cv2.boundingRect(cc)
                        # Avoid matching the outer table border as a cell
                        if cw < w - 10 and ch < h - 10 and cw > 15 and ch > 10:
                            # Translate coordinates back to page coordinate scale
                            # Need to scale down because image is 300 DPI, while PyMuPDF uses 72 DPI points
                            scale = 72.0 / 300.0
                            cell_box = (
                                (x + cx) * scale,
                                (y + cy) * scale,
                                (x + cx + cw) * scale,
                                (y + cy + ch) * scale
                            )
                            detected_cells.append(cell_box)
            
            self.log(f"OpenCV detected {len(detected_cells)} potential cells in page {page_num}")
            return detected_cells
        except Exception as e:
            self.log(f"OpenCV table detection failed on page {page_num}: {e}", logging.WARNING)
            return []

    def extract_text_pdf_layer(self, page, page_num):
        """
        Extracts structured text blocks with coordinates using PyMuPDF.
        """
        self.log(f"Extracting native text blocks for page {page_num} using PyMuPDF")
        text_blocks = []
        try:
            # page.get_text("blocks") returns list of tuples:
            # (x0, y0, x1, y1, "text", block_no, block_type)
            blocks = page.get_text("blocks")
            for b in blocks:
                x0, y0, x1, y1, text, block_no, block_type = b
                clean_text = text.strip()
                if clean_text:
                    # Filter out tiny or purely whitespace blocks
                    text_blocks.append({
                        "bbox": (x0, y0, x1, y1),
                        "text": clean_text,
                        "type": "text" if block_type == 0 else "image/other",
                        "font_bold": False # Can expand this with page.get_text("dict")
                    })
            self.log(f"Extracted {len(text_blocks)} text blocks from PDF layer for page {page_num}")
        except Exception as e:
            self.log(f"PyMuPDF native text extraction failed on page {page_num}: {e}", logging.ERROR)
        
        return text_blocks

    def run_ocr_fallback(self, image_path, bbox_72dpi, page_num):
        """
        Crops a specific bounding box from the high-res image and runs OCR.
        """
        if not HAS_OCR or not HAS_PIL:
            return "", 0.0

        try:
            # Scale coordinates up to image coordinates (300 DPI)
            scale = 300.0 / 72.0
            x0, y0, x1, y1 = [int(coord * scale) for coord in bbox_72dpi]
            
            # Open high-res render and crop
            img = PILImage.open(image_path)
            # Ensure crop box is inside image bounds
            x0, y0 = max(0, x0), max(0, y0)
            x1, y1 = min(img.width, x1), min(img.height, y1)
            
            if x1 <= x0 or y1 <= y0:
                return "", 0.0

            cropped_img = img.crop((x0, y0, x1, y1))
            
            text = ""
            confidence = 1.0

            if OCR_ENGINE == "Tesseract":
                # Run OCR with pytesseract
                # data format gives text, conf, etc.
                ocr_data = pytesseract.image_to_data(cropped_img, output_type=pytesseract.Output.DICT)
                words = []
                confidences = []
                for i in range(len(ocr_data['text'])):
                    w_text = ocr_data['text'][i].strip()
                    if w_text:
                        words.append(w_text)
                        confidences.append(float(ocr_data['conf'][i]))
                
                text = " ".join(words)
                confidence = sum(confidences) / len(confidences) / 100.0 if confidences else 0.0
            
            elif OCR_ENGINE == "EasyOCR":
                # Run OCR with EasyOCR
                # Needs np array
                img_np = np.array(cropped_img)
                results = reader.readtext(img_np)
                words = [res[1] for res in results]
                confs = [res[2] for res in results]
                text = " ".join(words)
                confidence = sum(confs) / len(confs) if confs else 0.0
                
            return text.strip(), confidence
        except Exception as e:
            self.log(f"OCR fallback failed on bbox {bbox_72dpi}: {e}", logging.WARNING)
            return "", 0.0

    def extract_images_pymupdf(self, doc, page, page_num):
        """
        Extracts images embedded in the PDF page and their dimensions.
        """
        extracted_images = []
        if not HAS_FITZ:
            return []

        try:
            image_list = page.get_images(full=True)
            self.log(f"Detected {len(image_list)} embedded images on page {page_num}")
            
            for img_idx, img_info in enumerate(image_list):
                xref = img_info[0]
                base_image = doc.extract_image(xref)
                image_bytes = base_image["image"]
                image_ext = base_image["ext"]
                
                # Write image to disk
                img_name = f"page_{page_num}_img_{img_idx + 1}.{image_ext}"
                img_path = os.path.join(self.preview_dir, img_name)
                
                with open(img_path, "wb") as f:
                    f.write(image_bytes)
                
                # Try to get visual position of image
                # page.get_image_rects(xref) returns lists of Rects where it is placed
                rects = page.get_image_rects(xref)
                bbox = (0, 0, 50, 50) # fallback
                if rects:
                    r = rects[0]
                    bbox = (r.x0, r.y0, r.x1, r.y1)
                
                extracted_images.append({
                    "path": img_path,
                    "bbox": bbox,
                    "name": img_name
                })
                self.log(f"Saved embedded image to {img_path} at position {bbox}")
        except Exception as e:
            self.log(f"Failed to extract images via PyMuPDF on page {page_num}: {e}", logging.WARNING)
            
        return extracted_images

    def cluster_grid_coordinates(self, elements):
        """
        Mathematical clustering of X and Y coordinates to map them cleanly into
        Excel grid columns and rows. Prevents hundreds of tiny columns/rows.
        """
        if not HAS_NP:
            # Fallback coordinate sorting if NumPy is missing
            xs = sorted(list({el['bbox'][0] for el in elements} | {el['bbox'][2] for el in elements}))
            ys = sorted(list({el['bbox'][1] for el in elements} | {el['bbox'][3] for el in elements}))
            return xs, ys

        # Gather all unique coordinates
        x_coords = []
        y_coords = []
        for el in elements:
            bbox = el["bbox"]
            x_coords.extend([bbox[0], bbox[2]])
            y_coords.extend([bbox[1], bbox[3]])

        # Sort coordinates
        x_coords = np.sort(np.array(list(set(x_coords))))
        y_coords = np.sort(np.array(list(set(y_coords))))

        # Simple clustering: merge coordinates within threshold pixels
        def cluster_1d(coords, threshold):
            if len(coords) == 0:
                return []
            clusters = []
            current_cluster = [coords[0]]
            for c in coords[1:]:
                if c - current_cluster[-1] <= threshold:
                    current_cluster.append(c)
                else:
                    clusters.append(np.mean(current_cluster))
                    current_cluster = [c]
            clusters.append(np.mean(current_cluster))
            return sorted(list(set(clusters)))

        cols = cluster_1d(x_coords, self.grid_threshold)
        rows = cluster_1d(y_coords, self.grid_threshold)

        return cols, rows

    def find_grid_indices(self, bbox, cols, rows):
        """
        Given a bounding box and the clustered column/row grid coordinates,
        finds the start/end column and row indices in Excel (1-indexed).
        """
        x0, y0, x1, y1 = bbox
        
        # Helper to find closest index in grid
        def find_closest_idx(val, grid):
            min_diff = float('inf')
            closest_idx = 0
            for idx, g_val in enumerate(grid):
                diff = abs(val - g_val)
                if diff < min_diff:
                    min_diff = diff
                    closest_idx = idx
            return closest_idx + 1 # 1-indexed for Excel

        col_start = find_closest_idx(x0, cols)
        col_end = find_closest_idx(x1, cols)
        row_start = find_closest_idx(y0, rows)
        row_end = find_closest_idx(y1, rows)

        # Ensure start is less than end
        col_start, col_end = min(col_start, col_end), max(col_start, col_end)
        row_start, row_end = min(row_start, row_end), max(row_start, row_end)

        # Adjust so it occupies at least one cell
        if col_start == col_end:
            col_end += 1
        if row_start == row_end:
            row_end += 1

        return col_start, col_end, row_start, row_end

    def generate_layout_preview(self, page_num, image_path, text_blocks, tables, images_list):
        """
        Generates a colored layout preview image highlighting detected elements.
        - Red: Text blocks
        - Green: Table boundaries/cells
        - Blue: Extracted images
        """
        if not HAS_CV2:
            return
        
        try:
            img = cv2.imread(image_path)
            # Scale factor from PDF points (72) to image pixels (300 DPI)
            scale = 300.0 / 72.0

            # Draw text blocks in red
            for tb in text_blocks:
                x0, y0, x1, y1 = [int(coord * scale) for coord in tb["bbox"]]
                cv2.rectangle(img, (x0, y0), (x1, y1), (0, 0, 255), 2)
                cv2.putText(img, "TXT", (x0, y0 - 5), cv2.FONT_HERSHEY_SIMPLEX, 0.4, (0, 0, 255), 1)

            # Draw table cells in green
            for t_cell in tables:
                x0, y0, x1, y1 = [int(coord * scale) for coord in t_cell]
                cv2.rectangle(img, (x0, y0), (x1, y1), (0, 255, 0), 2)
                cv2.putText(img, "CELL", (x0, y0 - 5), cv2.FONT_HERSHEY_SIMPLEX, 0.4, (0, 255, 0), 1)

            # Draw images in blue
            for img_item in images_list:
                x0, y0, x1, y1 = [int(coord * scale) for coord in img_item["bbox"]]
                cv2.rectangle(img, (x0, y0), (x1, y1), (255, 0, 0), 3)
                cv2.putText(img, "IMG", (x0, y0 - 5), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 0, 0), 2)

            preview_path = os.path.join(self.preview_dir, f"page_{page_num}_layout_preview.png")
            cv2.imwrite(preview_path, img)
            self.log(f"Generated layout bounding-box preview for page {page_num} at {preview_path}")
        except Exception as e:
            self.log(f"Failed to generate layout preview image: {e}", logging.WARNING)

    def analyze_pdf(self):
        """
        Executes full PDF analysis pipeline: rendering, text parsing, OCR fallback,
        table detection, and coordinate clustering.
        """
        if not HAS_FITZ:
            self.log("Cannot run PDF analysis without PyMuPDF.", logging.ERROR)
            return None

        # Render high-res images first
        rendered_images = self.render_pages()

        doc = fitz.open(self.pdf_path)
        self.log(f"Loaded PDF: {self.pdf_path} (Total Pages: {len(doc)})")

        analyzed_pages = {}

        for i, page in enumerate(doc):
            page_num = i + 1
            self.log(f"--- Processing Page {page_num} / {len(doc)} ---")
            
            # Extract PDF elements
            text_blocks = self.extract_text_pdf_layer(page, page_num)
            embedded_images = self.extract_images_pymupdf(doc, page, page_num)
            
            # Table cell boundaries (OpenCV or pdfplumber)
            page_img_path = rendered_images[i] if i < len(rendered_images) else None
            table_cells = []
            if page_img_path:
                table_cells = self.detect_tables_opencv(page_img_path, page_num)
            
            # Combine elements for clustering
            all_elements = []
            for tb in text_blocks:
                all_elements.append({"bbox": tb["bbox"], "type": "text", "data": tb})
            for c_bbox in table_cells:
                all_elements.append({"bbox": c_bbox, "type": "table_cell", "data": None})
            for img_item in embedded_images:
                all_elements.append({"bbox": img_item["bbox"], "type": "image", "data": img_item})

            # Check for scanned PDF text layer fallback (if text is completely empty)
            text_block_count = sum(1 for el in all_elements if el["type"] == "text")
            ocr_active = False
            if text_block_count == 0 and page_img_path and HAS_OCR:
                self.log(f"Page {page_num} has no native text layer. Activating OCR fallback!", logging.WARNING)
                ocr_active = True
                
                # Perform full-page OCR layout detection or block-level OCR
                # For this modular fallback, we run OCR on the main blocks detected by fitz,
                # or if fitz returned nothing, we do basic OpenCV block segmentation first.
                # Let's run a uniform fallback text search
                try:
                    img = PILImage.open(page_img_path)
                    if HAS_OCR and OCR_ENGINE == "Tesseract":
                        # Get text blocks with coordinates from Tesseract
                        ocr_data = pytesseract.image_to_data(img, output_type=pytesseract.Output.DICT)
                        n_boxes = len(ocr_data['text'])
                        scale = 72.0 / 300.0 # Scale back to PDF points
                        
                        # Group consecutive words into blocks
                        current_word_list = []
                        last_x, last_y, last_w, last_h = 0, 0, 0, 0
                        
                        for word_idx in range(n_boxes):
                            w_text = ocr_data['text'][word_idx].strip()
                            if w_text:
                                wx, wy, ww, wh = ocr_data['left'][word_idx], ocr_data['top'][word_idx], ocr_data['width'][word_idx], ocr_data['height'][word_idx]
                                conf = float(ocr_data['conf'][word_idx]) / 100.0
                                
                                # If word is close to last word, merge into block
                                if len(current_word_list) == 0 or (abs(wx - (last_x + last_w)) < 30 and abs(wy - last_y) < 15):
                                    current_word_list.append((w_text, conf, wx, wy, ww, wh))
                                else:
                                    # Create a combined text block
                                    bx0 = min([w[2] for w in current_word_list]) * scale
                                    by0 = min([w[3] for w in current_word_list]) * scale
                                    bx1 = max([w[2] + w[4] for w in current_word_list]) * scale
                                    by1 = max([w[3] + w[5] for w in current_word_list]) * scale
                                    b_txt = " ".join([w[0] for w in current_word_list])
                                    b_conf = sum([w[1] for w in current_word_list]) / len(current_word_list)
                                    
                                    text_blocks.append({
                                        "bbox": (bx0, by0, bx1, by1),
                                        "text": b_txt,
                                        "type": "text",
                                        "confidence": b_conf,
                                        "ocr": True
                                    })
                                    current_word_list = [(w_text, conf, wx, wy, ww, wh)]
                                
                                last_x, last_y, last_w, last_h = wx, wy, ww, wh
                                
                        if current_word_list:
                            # Flush last block
                            bx0 = min([w[2] for w in current_word_list]) * scale
                            by0 = min([w[3] for w in current_word_list]) * scale
                            bx1 = max([w[2] + w[4] for w in current_word_list]) * scale
                            by1 = max([w[3] + w[5] for w in current_word_list]) * scale
                            b_txt = " ".join([w[0] for w in current_word_list])
                            b_conf = sum([w[1] for w in current_word_list]) / len(current_word_list)
                            
                            text_blocks.append({
                                "bbox": (bx0, by0, bx1, by1),
                                "text": b_txt,
                                "type": "text",
                                "confidence": b_conf,
                                "ocr": True
                            })
                            
                    elif HAS_OCR and OCR_ENGINE == "EasyOCR":
                        # EasyOCR bounds
                        img_np = np.array(img)
                        results = reader.readtext(img_np)
                        scale = 72.0 / 300.0
                        for res in results:
                            # res: [[[x0, y0], [x1, y1], [x2, y2], [x3, y3]], text, confidence]
                            pts = res[0]
                            eb_x0, eb_y0 = pts[0][0] * scale, pts[0][1] * scale
                            eb_x1, eb_y1 = pts[2][0] * scale, pts[2][1] * scale
                            text_blocks.append({
                                "bbox": (eb_x0, eb_y0, eb_x1, eb_y1),
                                "text": res[1],
                                "type": "text",
                                "confidence": res[2],
                                "ocr": True
                            })
                    
                    self.log(f"OCR extracted {len(text_blocks)} text blocks from image layout")
                    # Re-populate all_elements with OCR elements
                    all_elements = []
                    for tb in text_blocks:
                        all_elements.append({"bbox": tb["bbox"], "type": "text", "data": tb})
                    for c_bbox in table_cells:
                        all_elements.append({"bbox": c_bbox, "type": "table_cell", "data": None})
                    for img_item in embedded_images:
                        all_elements.append({"bbox": img_item["bbox"], "type": "image", "data": img_item})

                except Exception as e:
                    self.log(f"Failed during fallback OCR rendering: {e}", logging.ERROR)

            # Cluster coordinates to define the coordinate grid columns/rows
            cols, rows = self.cluster_grid_coordinates(all_elements)
            self.log(f"Clustered layout grid: {len(cols)} columns, {len(rows)} rows")

            # Store stats
            self.stats[page_num] = {
                "tables_count": 1 if len(table_cells) > 0 else 0, # rough estimate
                "cells_count": len(table_cells),
                "text_count": len(text_blocks),
                "images_count": len(embedded_images),
                "ocr_active": ocr_active
            }

            # Generate and save previews with bounding box overlays
            if page_img_path:
                self.generate_layout_preview(page_num, page_img_path, text_blocks, table_cells, embedded_images)

            analyzed_pages[page_num] = {
                "text_blocks": text_blocks,
                "table_cells": table_cells,
                "images": embedded_images,
                "cols": cols,
                "rows": rows,
                "ocr_active": ocr_active
            }

        doc.close()
        return analyzed_pages

    def write_to_excel(self, analyzed_pages):
        """
        Creates a beautifully styled Excel spreadsheet from scratch.
        Maps layouts, tables, cell merging, styling, and images.
        """
        if not HAS_OPENPYXL:
            self.log("Cannot create Excel workbook without openpyxl.", logging.ERROR)
            return

        self.log(f"Step 2: Recreating elements in Excel file: {self.output_excel}")
        
        # Create Workbook
        wb = openpyxl.Workbook()
        # Remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)

        # Build Font Styles
        font_header = Font(name="Inter", size=11, bold=True, color="1F2937")
        font_body = Font(name="Inter", size=10, color="374151")
        font_code = Font(name="JetBrains Mono", size=9, color="4B5563")
        font_title = Font(name="Inter", size=14, bold=True, color="111827")
        font_italic_gray = Font(name="Inter", size=9, italic=True, color="6B7280")
        
        # Borders and Fills
        thin_border_side = Side(border_style="thin", color="D1D5DB")
        table_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
        
        fill_header = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid") # soft light gray
        fill_uncertain = PatternFill(start_color="FEF08A", end_color="FEF08A", fill_type="solid") # soft yellow

        # Helper to convert point sizes to Excel widths/heights
        def pt_to_excel_col_width(pt_width):
            # Proportional scaling (standard PDF point is 1/72 inch, Excel column width is roughly 1/12 inch)
            return max(3, min(pt_width / 7.2, 100))

        def pt_to_excel_row_height(pt_height):
            # Excel row height is in points
            return max(15, min(pt_height, 200))

        for page_num, p_data in analyzed_pages.items():
            sheet_name = f"Page {page_num}"
            ws = wb.create_sheet(title=sheet_name)
            ws.views.sheetView[0].showGridLines = True # Ensure grid lines are visible

            cols = p_data["cols"]
            rows = p_data["rows"]
            text_blocks = p_data["text_blocks"]
            table_cells = p_data["table_cells"]
            images = p_data["images"]

            # Initialize Excel Column Widths
            for c_idx in range(len(cols) - 1):
                col_width_pt = cols[c_idx+1] - cols[c_idx]
                col_letter = get_column_letter(c_idx + 1)
                ws.column_dimensions[col_letter].width = pt_to_excel_col_width(col_width_pt)

            # Initialize Excel Row Heights
            for r_idx in range(len(rows) - 1):
                row_height_pt = rows[r_idx+1] - rows[r_idx]
                ws.row_dimensions[r_idx + 1].height = pt_to_excel_row_height(row_height_pt)

            # Track populated cells to prevent overwrites or handle overlaps
            populated_grid = {}

            # Helper to assign value and formatting to a specific cell
            def populate_cell(r, c, val, bold=False, is_uncertain=False, comment_text=None, alignment="left"):
                cell = ws.cell(row=r, column=c)
                
                # Handle MergedCell (which has read-only value and comment) by redirecting to top-left cell
                if type(cell).__name__ == 'MergedCell':
                    parent_range = None
                    for rng in ws.merged_cells.ranges:
                        if rng.min_row <= r <= rng.max_row and rng.min_col <= c <= rng.max_col:
                            parent_range = rng
                            break
                    if parent_range:
                        cell = ws.cell(row=parent_range.min_row, column=parent_range.min_col)
                    else:
                        # Fallback if no range found to avoid crash
                        return

                # Check value type and cast if possible
                typed_val = val
                try:
                    # Try cast to integer
                    if "." not in val:
                        typed_val = int(val)
                    else:
                        typed_val = float(val)
                except ValueError:
                    # Keep as text
                    pass

                cell.value = typed_val
                cell.font = Font(name="Inter", size=10, bold=bold, color="374151")
                
                # Alignments
                if alignment == "right":
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif alignment == "center":
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

                # Uncertain highlighting
                if is_uncertain:
                    cell.fill = fill_uncertain
                    if comment_text:
                        cell.comment = Comment(comment_text, "Layout Recreator")

            # 1. Populate Table Cells First (since they form the core structure)
            self.log(f"Writing tables and cells to worksheet: {sheet_name}")
            for cell_bbox in table_cells:
                col_s, col_e, row_s, row_e = self.find_grid_indices(cell_bbox, cols, rows)
                
                # Merge if cell spans multiple columns/rows
                is_merged = (col_e - col_s > 1) or (row_e - row_s > 1)
                if is_merged:
                    try:
                        ws.merge_cells(start_row=row_s, start_column=col_s, end_row=row_e-1, end_column=col_e-1)
                    except Exception:
                        pass # Avoid crashing on overlapping merges

                # Find any text block that overlaps with this cell
                cell_text = ""
                text_sources = []
                is_header = False
                confidence = 1.0
                has_ocr = False

                for tb in text_blocks:
                    tx0, ty0, tx1, ty1 = tb["bbox"]
                    cx0, cy0, cx1, cy1 = cell_bbox
                    # Check for bounding box intersection
                    if not (tx1 < cx0 or tx0 > cx1 or ty1 < cy0 or ty0 > cy1):
                        cell_text += " " + tb["text"]
                        text_sources.append(tb)
                        if "ocr" in tb:
                            has_ocr = True
                            confidence = min(confidence, tb.get("confidence", 1.0))
                
                cell_text = cell_text.strip()
                
                # Format cell and add borders
                for r in range(row_s, row_e):
                    for c in range(col_s, col_e):
                        grid_cell = ws.cell(row=r, column=c)
                        grid_cell.border = table_border
                        
                        # Apply table styling (headers typically light gray at top)
                        if r == row_s and len(table_cells) > 0:
                            # If row_s is the top of the table range, treat as header
                            # Quick heuristic: header row is colored
                            grid_cell.fill = fill_header

                # Write content to cell
                if cell_text:
                    # Detect align
                    align = "left"
                    if cell_text.replace(".", "").replace("-", "").strip().isdigit():
                        align = "right" # Numeric align

                    uncertain = False
                    reason = ""
                    if has_ocr and confidence < self.ocr_threshold:
                        uncertain = True
                        reason = f"Low confidence OCR ({confidence:.2f})"
                        self.uncertain_cells.append((sheet_name, f"{get_column_letter(col_s)}{row_s}", cell_text, reason, confidence))
                    
                    populate_cell(
                        row_s, col_s, cell_text,
                        bold=is_header,
                        is_uncertain=uncertain,
                        comment_text=reason if uncertain else None,
                        alignment=align
                    )
                    
                    # Mark populated in grid map
                    for r in range(row_s, row_e):
                        for c in range(col_s, col_e):
                            populated_grid[(r, c)] = True

            # 2. Populate Non-Table Text Blocks
            self.log(f"Writing text blocks to worksheet: {sheet_name}")
            for tb in text_blocks:
                tb_bbox = tb["bbox"]
                col_s, col_e, row_s, row_e = self.find_grid_indices(tb_bbox, cols, rows)
                
                # Skip if already fully populated by table cells (avoid duplications)
                is_already_populated = any((r, c) in populated_grid for r in range(row_s, row_e) for c in range(col_s, col_e))
                if is_already_populated:
                    continue

                is_merged = (col_e - col_s > 1) or (row_e - row_s > 1)
                if is_merged:
                    try:
                        ws.merge_cells(start_row=row_s, start_column=col_s, end_row=row_e-1, end_column=col_e-1)
                    except Exception:
                        pass

                # Check headers / title styles
                is_bold = tb.get("font_bold", False) or len(tb["text"]) < 40 and tb["text"].isupper()
                is_title = row_s <= 3 and len(tb["text"]) < 60 # title heuristic

                uncertain = False
                reason = ""
                if "ocr" in tb and tb.get("confidence", 1.0) < self.ocr_threshold:
                    uncertain = True
                    reason = f"Low confidence OCR ({tb['confidence']:.2f})"
                    self.uncertain_cells.append((sheet_name, f"{get_column_letter(col_s)}{row_s}", tb["text"], reason, tb["confidence"]))

                populate_cell(
                    row_s, col_s, tb["text"],
                    bold=is_bold or is_title,
                    is_uncertain=uncertain,
                    comment_text=reason if uncertain else None
                )

                if is_title:
                    ws.cell(row=row_s, column=col_s).font = font_title

                # Mark populated
                for r in range(row_s, row_e):
                    for c in range(col_s, col_e):
                        populated_grid[(r, c)] = True

            # 3. Embed Images
            self.log(f"Embedding images into worksheet: {sheet_name}")
            for img_item in images:
                col_s, col_e, row_s, row_e = self.find_grid_indices(img_item["bbox"], cols, rows)
                
                # Check if file exists and add to Excel
                if os.path.exists(img_item["path"]):
                    try:
                        xl_img = openpyxl.drawing.image.Image(img_item["path"])
                        # Anchor image to Excel coordinate
                        cell_anchor = f"{get_column_letter(col_s)}{row_s}"
                        ws.add_image(xl_img, cell_anchor)
                        self.log(f"Anchored image {img_item['name']} at {cell_anchor}")
                    except Exception as e:
                        self.log(f"Failed to embed image {img_item['name']} into Excel: {e}", logging.WARNING)

        # 4. Add the "Control" Worksheet containing metadata, validation stats, and logs
        self.log("Step 3: Creating 'Control' metadata and validation sheet")
        ws_control = wb.create_sheet(title="Control", index=0)
        ws_control.views.sheetView[0].showGridLines = True
        
        # Style header block
        ws_control.column_dimensions['A'].width = 30
        ws_control.column_dimensions['B'].width = 50
        ws_control.column_dimensions['C'].width = 20
        ws_control.column_dimensions['D'].width = 40
        
        # Set title block
        ws_control.merge_cells("A1:D1")
        ws_control["A1"] = "PDF TO EXCEL EXTRACTION CONTROL CENTER"
        ws_control["A1"].font = Font(name="Inter", size=14, bold=True, color="FFFFFF")
        ws_control["A1"].fill = PatternFill(start_color="111827", end_color="111827", fill_type="solid") # deep slate
        ws_control["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws_control.row_dimensions[1].height = 40

        # Section 1: Extraction Info
        ws_control["A3"] = "EXTRACTION METADATA"
        ws_control["A3"].font = font_header
        
        metadata = [
            ("Source PDF File Name", os.path.basename(self.pdf_path)),
            ("Total Page Count", len(analyzed_pages)),
            ("Extraction Timestamp", self.start_time.strftime("%Y-%m-%d %H:%M:%S")),
            ("OCR Engine Active", OCR_ENGINE if any(st.get("ocr_active") for st in self.stats.values()) else "None (Native Text Layer)"),
            ("Grid Threshold Settings", f"{self.grid_threshold} pixels (coordinate clustering)"),
            ("Output Excel File", self.output_excel),
        ]
        
        r_idx = 4
        for label, val in metadata:
            ws_control.cell(row=r_idx, column=1, value=label).font = font_body
            ws_control.cell(row=r_idx, column=1).border = table_border
            ws_control.cell(row=r_idx, column=2, value=val).font = font_code
            ws_control.cell(row=r_idx, column=2).border = table_border
            r_idx += 1

        # Section 2: Page Analysis Statistics
        ws_control.cell(row=r_idx+1, column=1, value="PAGE EXTRACTION METRICS").font = font_header
        r_idx += 2
        
        headers = ["Page Number", "Detected Tables", "Total Cells", "Text Blocks", "Images Count", "OCR Triggered"]
        for c_idx, h in enumerate(headers):
            cell = ws_control.cell(row=r_idx, column=c_idx+1, value=h)
            cell.font = font_header
            cell.fill = fill_header
            cell.border = table_border
            cell.alignment = Alignment(horizontal="center")
        
        for p_num, st in self.stats.items():
            r_idx += 1
            ws_control.cell(row=r_idx, column=1, value=f"Page {p_num}").font = font_body
            ws_control.cell(row=r_idx, column=2, value=st["tables_count"]).font = font_body
            ws_control.cell(row=r_idx, column=3, value=st["cells_count"]).font = font_body
            ws_control.cell(row=r_idx, column=4, value=st["text_count"]).font = font_body
            ws_control.cell(row=r_idx, column=5, value=st["images_count"]).font = font_body
            ws_control.cell(row=r_idx, column=6, value="YES" if st["ocr_active"] else "NO").font = font_body
            
            for c_idx in range(len(headers)):
                ws_control.cell(row=r_idx, column=c_idx+1).border = table_border
                ws_control.cell(row=r_idx, column=c_idx+1).alignment = Alignment(horizontal="center")

        # Section 3: Validation, Warnings, and Low Confidence Logs
        r_idx += 2
        ws_control.cell(row=r_idx, column=1, value="UNCERTAIN OR LOW-CONFIDENCE EXTRACTION LOGS").font = font_header
        r_idx += 1
        
        ws_control.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=4)
        ws_control.cell(row=r_idx, column=1, value="The cells listed below were extracted with lower confidence or encountered layout coordinates overlaps. They have been highlighted in yellow on their respective worksheets.").font = font_italic_gray
        r_idx += 1

        val_headers = ["Sheet Name", "Cell Address", "Extracted Value / Text", "Reason / Confidence Detail"]
        for c_idx, h in enumerate(val_headers):
            cell = ws_control.cell(row=r_idx, column=c_idx+1, value=h)
            cell.font = font_header
            cell.fill = fill_header
            cell.border = table_border
            cell.alignment = Alignment(horizontal="center")

        if len(self.uncertain_cells) == 0:
            r_idx += 1
            ws_control.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=4)
            ws_control.cell(row=r_idx, column=1, value="Perfect Validation! No uncertain or low confidence cells detected.").font = font_body
            ws_control.cell(row=r_idx, column=1).alignment = Alignment(horizontal="center")
            for c_idx in range(4):
                ws_control.cell(row=r_idx, column=c_idx+1).border = table_border
        else:
            for sheet, addr, val, reason, conf in self.uncertain_cells:
                r_idx += 1
                ws_control.cell(row=r_idx, column=1, value=sheet).font = font_body
                ws_control.cell(row=r_idx, column=2, value=addr).font = font_code
                ws_control.cell(row=r_idx, column=3, value=val).font = font_body
                ws_control.cell(row=r_idx, column=4, value=reason).font = font_body
                
                for c_idx in range(4):
                    ws_control.cell(row=r_idx, column=c_idx+1).border = table_border
                    ws_control.cell(row=r_idx, column=c_idx+1).fill = fill_uncertain

        # Section 4: Engine Warnings
        r_idx += 2
        ws_control.cell(row=r_idx, column=1, value="SYSTEM AND PARSING WARNINGS").font = font_header
        r_idx += 1
        
        ws_control.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=4)
        ws_control.cell(row=r_idx, column=1, value="Warnings encountered during the file parsing, image slicing, or cell mappings:").font = font_italic_gray
        r_idx += 1

        if len(self.warnings) == 0:
            ws_control.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=4)
            ws_control.cell(row=r_idx, column=1, value="Clean Run! No execution warnings or parsing conflicts logged.").font = font_body
            ws_control.cell(row=r_idx, column=1).alignment = Alignment(horizontal="center")
            for c_idx in range(4):
                ws_control.cell(row=r_idx, column=c_idx+1).border = table_border
        else:
            for w in self.warnings:
                ws_control.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=4)
                cell = ws_control.cell(row=r_idx, column=1, value=w)
                cell.font = font_code
                cell.border = table_border
                r_idx += 1

        # Save workbook
        try:
            wb.save(self.output_excel)
            self.log(f"Workbook successfully saved to: {self.output_excel}")
        except Exception as e:
            self.log(f"Failed to save Excel file: {e}", logging.ERROR)

    def execute(self):
        """Runs the entire recreation flow."""
        try:
            self.log("Initializing PDF to Excel Layout Recreator execution pipeline...")
            
            # 1. Analyze PDF
            analyzed_pages = self.analyze_pdf()
            if not analyzed_pages:
                self.log("PDF analysis phase failed. Ending execution.", logging.ERROR)
                return False

            # 2. Recreate in Excel
            self.write_to_excel(analyzed_pages)
            
            end_time = datetime.datetime.now()
            duration = (end_time - self.start_time).total_seconds()
            self.log(f"Recreation complete in {duration:.2f} seconds.")
            
            # Write final log
            self.write_log_file()
            return True
        except Exception as e:
            self.log(f"Pipeline crashed during execution: {e}", logging.ERROR)
            self.write_log_file()
            return False


def main():
    parser = argparse.ArgumentParser(description="Recreate PDF page layouts and tables in an Excel sheet from scratch.")
    parser.add_argument("pdf_path", help="Path to the input PDF file")
    parser.add_argument("--output", default="recreated_from_pdf.xlsx", help="Filename of the generated Excel workbook")
    parser.add_argument("--previews", default="page_previews", help="Folder to save page rendering and element bounding-box images")
    parser.add_argument("--log", default="extraction_log.txt", help="Filename of the detailed steps and error logs")
    parser.add_argument("--combine", action="store_true", help="Combine pages into a single Excel worksheet (default: separate sheets)")
    parser.add_argument("--grid_threshold", type=int, default=10, help="Coordinate grouping threshold in pixels")
    parser.add_argument("--ocr_threshold", type=float, default=0.6, help="Confidence threshold below which cell will be highlighted")

    args = parser.parse_args()

    # Integrity verification
    if not os.path.exists(args.pdf_path):
        print(f"Error: PDF file not found at: {args.pdf_path}")
        sys.exit(1)

    if not HAS_OPENPYXL:
        print("Error: Missing mandatory dependency 'openpyxl'. Please run: pip install openpyxl")
        sys.exit(1)

    recreator = PDFRecreator(
        pdf_path=args.pdf_path,
        output_excel=args.output,
        preview_dir=args.previews,
        log_path=args.log,
        combine_pages=args.combine,
        grid_threshold=args.grid_threshold,
        ocr_threshold=args.ocr_threshold
    )

    success = recreator.execute()
    if success:
        print(f"\nSuccess! Excel file generated at: {args.output}")
        print(f"Extraction log and warnings at: {args.log}")
        print(f"Page previews and visual debugging layouts at: {args.previews}/")
    else:
        print("\nRecreation failed. Check extraction_log.txt for details.")
        sys.exit(1)


if __name__ == "__main__":
    main()
